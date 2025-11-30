import andes
import numpy as np 
import pandas as pd

def ANDES_Case_Runner(case, T_END):
    andes.config_logger(stream_level=50) # Logging verbosity: 50 = critial only
    ssys = andes.load(case)

    # Run power flow and dynamic sim
    ssys.PQ.config.p2p = 1.0
    ssys.PQ.config.p2i = 0
    ssys.PQ.config.p2z = 0

    ssys.PQ.config.q2q = 1.0
    ssys.PQ.config.q2i = 0
    ssys.PQ.config.q2z = 0

    ssys.PFlow.run()
    ssys.TDS.config.tf = T_END
    ssys.TDS.config.tstep = 0.02
    print('Running ANDES dynamic simulation...')
    ssys.TDS.run()
    print('Simulation finished')

    ssys.TDS.load_plotter()

    # fig, ax = ssys.TDS.plt.plot(ssys.GENROU.omega)
    # fig.savefig("IEEE14_omega_plot.png", dpi=300)

    # fig, ax = ssys.TDS.plt.plot(ssys.GENROU.v)
    # fig.savefig("IEEE14_v_plot.png", dpi=300)

    print("Load active power at end of sim: ", ssys.PQ.Ppf.v)
    ssys_csv = ssys.TDS.plt.export_csv()
    return ssys_csv

import openpyxl

def append_alter_rows(file_path, u_opt, time):
    """
    Append rows to the 'Alter' sheet based on u_opt and time.

    Parameters
    ----------
    file_path : str
        Path to the Excel file to modify.
    u_opt : numpy array of shape (3,)
        Three control values (but only their index is used for dev).
    time : float
        Time value to put in column 't'.
    """

    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Alter"]

    # Determine next uid / idx / name index
    # Skip header row (assumed row 1)
    last_row = sheet.max_row

    if last_row == 1:
        next_uid = 0
        next_idx = 1
        name_counter = 1
    else:
        # Read existing values from the last row
        next_uid = sheet.cell(row=last_row, column=1).value + 1
        next_idx = sheet.cell(row=last_row, column=2).value + 1
        
        # Extract name numbers: "Alter_1"
        last_name = sheet.cell(row=last_row, column=4).value
        name_counter = int(last_name.split("_")[1]) + 1
    
    dev_map = ["PQ_1", "PQ_2", "PQ_10"]

    # Add a new row for each element in u_opt
    for i in range(len(u_opt)):
        
        row = [
            next_uid + i,          # uid
            next_idx + i,          # idx
            1,                     # u
            f"Alter_{name_counter + i}",  # name
            time,                  # t
            "PQ",                  # model
            dev_map[i],           # dev → PQ_1, PQ_2, PQ_3
            "Ppf",                 # src (blank unless you want otherwise)
            "v",                   # attr
            "=",                   # method
            u_opt[i],              # amount
            0,                     # rand
            0,                     # lb
            0                      # ub
        ]

        sheet.append(row)

    wb.save(file_path)


def data_cleaner(data):
    """
    Filters the input data based on time segments to retain multiples of 0.02s.
    Preserves the input format (Pandas DataFrame or NumPy Array) and all columns.
    
    Parameters:
    data: Input data (DataFrame or 2D Array). 
          The first column (index 0) MUST be the time column.
    
    Returns:
    Filtered data in the same format as the input.
    """
    # 1. Extract the time column for logic, handling both DataFrame and Array
    if isinstance(data, pd.DataFrame):
        t = data.iloc[:, 0].values # Convert col 0 to numpy array for speed
    elif hasattr(data, 'ndim') and data.ndim > 1:
        t = data[:, 0]             # Use first column of numpy array
    else:
        t = data                   # Fallback for 1D input
        
    # 2. Define masks for each segment
    # (Handling triplets at boundaries by assigning them to the 'Keep 3rd' groups)
    
    # 0s - 5s (Keep 3rd): Includes 5.00 triplet
    mask_0_5   = (t <= 5.01)
    
    # 5s - 6s (Keep All): Strictly between 5.00 and 6.00 triplets
    mask_5_6   = (t > 5.01) & (t < 5.99)
    
    # 6s - 11s (Keep 3rd): Includes 6.00 and 11.00 triplets
    mask_6_11  = (t >= 5.99) & (t < 11.01)
    
    # 11s - 12s (Keep All): Strictly between 11.00 and 12.00 triplets
    mask_11_12 = (t > 11.01) & (t < 11.99)
    
    # 12s - 17s (Keep 3rd): Includes 12.00 and 17.00 triplets
    mask_12_17 = (t >= 11.99) & (t < 17.01)
    
    # 17s - 20s (Keep All): After 17.00 triplet
    mask_17_20 = (t > 17.01)

    # 3. Get indices for each segment
    idx_0_5   = np.where(mask_0_5)[0]
    idx_5_6   = np.where(mask_5_6)[0]
    idx_6_11  = np.where(mask_6_11)[0]
    idx_11_12 = np.where(mask_11_12)[0]
    idx_12_17 = np.where(mask_12_17)[0]
    idx_17_20 = np.where(mask_17_20)[0]

    # 4. Apply filters (Keep every 3rd vs Keep all)
    sel_0_5   = idx_0_5[::3]
    sel_5_6   = idx_5_6
    sel_6_11  = idx_6_11[::3]
    sel_11_12 = idx_11_12
    sel_12_17 = idx_12_17[::3]
    sel_17_20 = idx_17_20

    # 5. Combine indices
    all_indices = np.concatenate([
        sel_0_5, sel_5_6, sel_6_11, sel_11_12, sel_12_17, sel_17_20
    ])
    all_indices.sort()

    # 6. Return the result in the ORIGINAL format
    if isinstance(data, pd.DataFrame):
        return data.iloc[all_indices]  # Return DataFrame with all columns
    elif data.ndim == 2:
        return data[all_indices, :]    # Return 2D Array with all columns
    else:
        return data[all_indices]       # Return 1D Array


def input_signal(P0, A, t, t0, tf):
    # # 1. Extract time column
    # if isinstance(data, pd.DataFrame):
    #     t = data.iloc[:, 0].values
    # elif hasattr(data, 'ndim') and data.ndim > 1:
    #     t = data[:, 0]
    # else:
    #     t = np.array(data)

    values = np.zeros_like(t)

    # 2. Define Masks
    mask_before = t < t0
    mask_after = t > tf
    mask_between = (t >= t0) & (t <= tf)

    # 3. Apply "Before" Logic
    values[mask_before] = P0

    # 4. Apply "After" Logic
    # Holds the value of sin(tf)
    values[mask_after] = A*np.sin(2*np.pi  * 0.5 * (tf-t0)**2 / 5)

    # 5. Apply "Between" Logic (Step Function)
    if np.any(mask_between):
        t_subset = t[mask_between]
        step = 0.02
        tolerance = 1e-5
        
        # Find nearest multiple of 0.02
        k = np.round(t_subset / step)
        t_nearest = k * step
        
        # Determine effective time (hold previous if not close enough to next)
        effective_t = np.where(t_subset >= t_nearest - tolerance, t_nearest, t_nearest - step)
        
        values[mask_between] = A*np.sin(2*np.pi  * 0.5 * (effective_t-t0)**2 / 5)

    return values